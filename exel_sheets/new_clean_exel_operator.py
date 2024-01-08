"""
Модуль для создания xlsx файла с данными
"""
import json
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

logging.basicConfig(level=logging.INFO)


def book_creater(data: dict, start_row: int = 2):
    """
    Основная функция для поиска по дате, которая создает объект exel, наполняет его данными и сохраняет в файл.
    На вход получает словарь с данными и значение стартовой строки, равняющеюся 2 по дефолту
    """

    # инициализация объекта openxl
    wb = Workbook()
    sheet = wb.active
    row = start_row

    # добавим строку с поисковой датой
    create_date_row(data=data, row=row, sheet=sheet, start=True)

    # добавим строку с заголовками
    create_headers(row, sheet=sheet)

    # заполняем данными
    sheet_filler(data, row=row, sheet=sheet)

    # сохраняет файл
    wb.save(f'results/result_{data["date"]}.xlsx')


def book_creator_multiple_dates(event_list: list, search: str, date_interval: str, type_search: str,
                                country: str = None, city: str = None):
    """
    Основная функция для создания объекта exel, когда используется поиск событий по фразе в диапазоне дат.

    :param event_list: список с событиями;
    :param search: поисковая фраза, используется для именования файла xlsx;
    :param date_interval: интервал с датами, используется для именования файла xlsx;
    :param type_search: тип поиска, используется для именования файла xlsx;
    :param country: страна или страны через пробел, используется для именования файла xlsx;
    :param city: город или города через пробел, используется для именования файла xlsx
    """
    wb = Workbook()
    sheet = wb.active
    row = 4
    search_header = sheet.cell(row=1, column=1, value="Поисковый запрос: ")
    search_header.font = Font(size=14, bold=True)
    search_header.alignment = Alignment(horizontal='right', vertical='center')
    search_header.fill = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')
    search_phrase = sheet.cell(row=1, column=2, value=search)
    search_phrase.font = Font(size=14, bold=True)
    search_phrase.alignment = Alignment(horizontal="left", vertical='center')
    search_phrase.fill = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')

    # добавим строку с поисковой датой
    create_date_row(data=event_list[0], row=row, sheet=sheet, start=True)
    logging.info(f"working with: {event_list[0]['date']}")

    # добавим строку с заголовками
    create_headers(row, sheet=sheet)

    # заполняем данными
    sheet_filler(event_list[0], row=row, sheet=sheet)

    start_shift = measure_json(event_list[0]) + 8
    row = start_shift

    for event in event_list[1:]:
        logging.info(f"working with: {event['date']}")
        print(f"started row is {row}")
        # добавим строку с поисковой датой
        create_date_row(data=event, row=row, sheet=sheet)

        # добавим строку с заголовками
        create_headers(row, sheet=sheet)

        # заполняем данными
        sheet_filler(event, row=row, sheet=sheet)

        row += measure_json(event) + 4
    # Если были переданы геопараметры, такие как страна или город, добавляем их в название файла.
    if all((country, city)):
        # сохраняет файл
        geo = sum([country.split(), city.split()], [])
    elif country:
        geo = country.split()
    elif city:
        geo = city.split()
    else:
        geo = None
    if type_search == "1":
        strict = "Строгий_поиск"
    elif type_search == "2":
        strict = "Неточный_поиск"
    if geo:
        filename = f'results/result_{strict}_{search}_{"_".join(geo)}_{date_interval}.xlsx'
    else:
        filename = f'results/result_{strict}_{search}_{date_interval}.xlsx'
    wb.save(filename)


def sheet_filler(data: dict, row: int, sheet: Workbook.active):
    """
    Данная функция забирает данные из json и наполняет соответствующие столбцы
    :param data: данные для заполнения
    :param row: текущая активная строка
    :param sheet: рабочий лист Excel
    """

    # наполнение данных для calend.ru
    holidays = enumerate(data["праздники (данные с calend.ru)"], row + 2)
    for val in holidays:
        sheet.cell(row=val[0], column=1, value=val[1])
        sheet[f"A{val[0]}"].alignment = Alignment(wrap_text=True)

    # наполнение данных по праздничным дням из вики
    international = enumerate(data["Праздничные события (данные Wiki)"]["международные"], row + 2)
    for val in international:
        sheet.cell(row=val[0], column=2, value=val[1])
        sheet[f"B{val[0]}"].alignment = Alignment(wrap_text=True)
    national = enumerate(data["Праздничные события (данные Wiki)"]["национальные"], row + 2)
    for val in national:
        sheet.cell(row=val[0], column=3, value=val[1])
        sheet[f"C{val[0]}"].alignment = Alignment(wrap_text=True)
    regional = enumerate(data["Праздничные события (данные Wiki)"]["региональные"], row + 2)
    for val in regional:
        sheet.cell(row=val[0], column=4, value=val[1])
        sheet[f"D{val[0]}"].alignment = Alignment(wrap_text=True)
    prof = enumerate(data["Праздничные события (данные Wiki)"]["профессиональные"], row + 2)
    for val in prof:
        sheet.cell(row=val[0], column=5, value=val[1])
        sheet[f"E{val[0]}"].alignment = Alignment(wrap_text=True)

    # наполнение событиями из вики
    if isinstance(data["события в этот день (данные Wiki)"], str):
        sheet.cell(row=row + 2, column=6, value=data["события в этот день (данные Wiki)"]).alignment = Alignment(
            wrap_text=True)
        # sheet[f"F{row+2}"].alignment = Alignment(wrap_text=True)

    elif isinstance(data["события в этот день (данные Wiki)"], list):
        events = enumerate(data["события в этот день (данные Wiki)"], row + 2)
        for val in events:
            sheet.cell(row=val[0], column=6, value=val[1])
            sheet[f"F{val[0]}"].alignment = Alignment(wrap_text=True)

    # События за xx век
    i = row + 2
    while i < row + 2 + len(data["События за XX век"].items()):
        for year, event in data["События за XX век"].items():
            sheet.cell(row=i, column=7, value=int(year))
            sheet.cell(row=i, column=8, value=event[0])
            sheet[f"H{i}"].alignment = Alignment(wrap_text=True)
            i += 1

    # События за xxi век
    i = row + 2
    while i < row + 2 + len(data["События за XXI век"].items()):
        for year, event in data["События за XXI век"].items():
            sheet.cell(row=i, column=9, value=int(year))
            sheet.cell(row=i, column=10, value=event[0])
            sheet[f"J{i}"].alignment = Alignment(wrap_text=True)
            i += 1

    # Затмения
    sheet.cell(row=row + 2, column=11,
               value=f"Ближайшее солнечное затмение: {data['Затмения']['closest_solar_eclipse']}")
    sheet.cell(row=row + 2, column=12, value=data["Затмения"]["days_to_solar_eclipse"])
    sheet.cell(row=row + 3, column=11, value=f"Ближайшее лунное затмение: {data['Затмения']['closest_lunar_eclipse']}")
    sheet.cell(row=row + 3, column=12, value=data["Затмения"]["days_to_lunar_eclipse"])

    # Астероиды
    j = row + 2
    if data['астероиды']:
        for aster in data['астероиды']:
            sheet.cell(row=j, column=13, value=aster['des'])
            sheet.cell(row=j, column=14, value=aster['cd'])
            sheet.cell(row=j, column=15, value=aster['dist'])
            sheet.cell(row=j, column=16, value=aster['dist_min'])
            sheet.cell(row=j, column=17, value=aster['dist_max'])
            j += 1


def styling(cell):
    """
    Данная функция добавляет стилей заголовкам
    :param cell: активная ячейка, к которой будут применены стили
    """
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=14, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='205DAC', end_color='205DAC', fill_type='solid')


def create_date_row(data: dict, row, sheet, start=False):
    """
    Данная фунция добавляет поисковую строку с датой
    :param data: словарь с данными для заполнения excel файла;
    :param row: активная строка;
    :param sheet: активный лист;
    :param start определяет, является ли заголовки стартовыми или нет:
    """
    if start:
        row = row + 1

    date = sheet.cell(row=row - 2, column=1, value="Дата: ")
    date.font = Font(size=14, bold=True)
    date.alignment = Alignment(horizontal="right", vertical='center')
    value = sheet.cell(row=row - 2, column=2, value=data["date"])
    value.font = Font(size=14, bold=True)
    value.alignment = Alignment(horizontal="left", vertical="center")
    if not start:
        delta_text = sheet.cell(row=row - 1, column=1, value="Количество дней до предыдущей даты:")
        delta_text.font = Font(size=14, bold=True)
        delta_text.alignment = Alignment(horizontal="right", vertical='center')
        delta_value = sheet.cell(row=row - 1, column=2, value=data['timedelta'])
        delta_value.font = Font(size=14, bold=True)
        delta_value.alignment = Alignment(horizontal="left", vertical='center')


def create_headers(row: int, sheet) -> None:
    """
    Данная функция создает заголовки для последующего наполнения данными из json
    :param row: активная строка;
    :param sheet: активный рабочий лист Excel
    """
    # праздники (данные с calend.ru)
    h1 = sheet.cell(row=row, column=1, value="праздники (данные с calend.ru)")
    styling(h1)
    sheet.merge_cells(f"A{row}:A{row + 1}")

    # Праздничные события (данные Wiki) со всеми подзаголовками
    h2 = sheet.cell(row=row, column=2, value="Праздничные события (данные Wiki)")
    styling(h2)
    sheet.merge_cells(f"B{row}:E{row}")
    sub_h2_1 = sheet.cell(row=row + 1, column=2, value="международные")
    styling(sub_h2_1)
    sub_h2_2 = sheet.cell(row=row + 1, column=3, value="национальные")
    styling(sub_h2_2)
    sub_h2_3 = sheet.cell(row=row + 1, column=4, value="региональные")
    styling(sub_h2_3)
    sub_h2_4 = sheet.cell(row=row + 1, column=5, value="профессиональные")
    styling(sub_h2_4)

    # события в этот день (данные Wiki)
    h3 = sheet.cell(row=row, column=6, value="события в этот день (данные Wiki)")
    styling(h3)
    sheet.merge_cells(f"F{row}:F{row + 1}")

    # События за XX век с подзаголовками год и событие
    h4 = sheet.cell(row=row, column=7, value="События за XX век")
    styling(h4)
    sheet.merge_cells(f"G{row}:H{row}")
    sub_h4_1 = sheet.cell(row=row + 1, column=7, value="Год")
    styling(sub_h4_1)
    sub_h4_2 = sheet.cell(row=row + 1, column=8, value="Событие")
    styling(sub_h4_2)

    # События за XXШ век с подзаголовками год и событие
    h5 = sheet.cell(row=row, column=9, value="События за XXI век")
    styling(h5)
    sheet.merge_cells(f"I{row}:J{row}")
    sub_h5_1 = sheet.cell(row=row + 1, column=9, value="Год")
    styling(sub_h5_1)
    sub_h5_2 = sheet.cell(row=row + 1, column=10, value="Событие")
    styling(sub_h5_2)

    # затмения
    h6 = sheet.cell(row=row, column=11, value="Затмения относительно указанной даты")
    styling(h6)
    sheet.merge_cells(f"K{row}:L{row}")
    sub_h6_1 = sheet.cell(row=row + 1, column=11, value="Дата")
    styling(sub_h6_1)
    sub_h6_2 = sheet.cell(row=row + 1, column=12, value="Количество дней до него")
    styling(sub_h6_2)
    sub_h6_2.alignment = Alignment(wrap_text=True, horizontal='center')

    # астероиды
    h7 = sheet.cell(row=row, column=13, value="Астероиды относительно указанной даты")
    styling(h7)
    sheet.merge_cells(f"M{row}:Q{row}")
    sub_h7_1 = sheet.cell(row=row + 1, column=13, value="des")
    styling(sub_h7_1)
    sub_h7_2 = sheet.cell(row=row + 1, column=14, value="cd")
    styling(sub_h7_2)
    sub_h7_3 = sheet.cell(row=row + 1, column=15, value="dist")
    styling(sub_h7_3)
    sub_h7_4 = sheet.cell(row=row + 1, column=16, value="dist_min")
    styling(sub_h7_4)
    sub_h7_5 = sheet.cell(row=row + 1, column=17, value="dist_max")
    styling(sub_h7_5)

    # Задаем ширину столбцов
    sheet.column_dimensions["A"].width = 60
    for col in ["B", "C", "D", "E"]:
        sheet.column_dimensions[f"{col}"].width = 30
    sheet.column_dimensions["F"].width = 50
    sheet.column_dimensions["G"].width = 5
    sheet.column_dimensions["H"].width = 65
    sheet.column_dimensions["I"].width = 5
    sheet.column_dimensions["J"].width = 65
    sheet.column_dimensions["K"].width = 40
    sheet.column_dimensions["L"].width = 14
    sheet.column_dimensions["M"].width = 10
    sheet.column_dimensions["N"].width = 16
    sheet.column_dimensions["O"].width = 19
    sheet.column_dimensions["P"].width = 19
    sheet.column_dimensions["Q"].width = 18


def measure_json(data: dict) -> int:
    """
    Данная функция измереят длинну данных в json, чтобы можно было посчитать, сколько строк эти данные займут
    :param data: словарь с данными для заполнения
    """

    largest = max(1, 2, max(len(data['Праздничные события (данные Wiki)']['международные']),
                            len(data['Праздничные события (данные Wiki)']['региональные']),
                            len(data['Праздничные события (данные Wiki)']['национальные']),
                            len(data['Праздничные события (данные Wiki)']['профессиональные'])),
                  len(data['События за XX век']), len(data['События за XXI век']), len(data['астероиды']),
                  len(data['праздники (данные с calend.ru)']), len(data['события в этот день (данные Wiki)']))
    logging.info(f"the lenght of {data['date']} is: {largest}")
    return largest


if __name__ == '__main__':
    with open("../results/result_10 января 2009.json") as f:
        data = json.load(f)

    book_creater(data=data)
